"""Non-executing DOCX and spreadsheet extraction."""

from __future__ import annotations

from pathlib import Path
import zipfile

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import base as limits
from .base import (
    Extraction,
    ExtractionError,
    FragmentCollector,
    enforce_extraction_budget,
    registry,
    snapshot_file,
)


def _validate_office_zip(path: Path) -> None:
    """Validate Office ZIP metadata without reading or expanding any member."""
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
    except Exception as exc:
        raise ExtractionError(f"failed to inspect Office ZIP file: {path}") from exc

    if len(members) > limits.MAX_ZIP_MEMBERS:
        raise ExtractionError(
            f"Office ZIP member count exceeds budget of {limits.MAX_ZIP_MEMBERS}"
        )
    total_uncompressed = 0
    for member in members:
        total_uncompressed += member.file_size
        if total_uncompressed > limits.MAX_ZIP_UNCOMPRESSED_BYTES:
            raise ExtractionError("Office ZIP total uncompressed size exceeds 100 MiB budget")
        if (
            member.filename.lower().endswith((".xml", ".rels"))
            and member.file_size > limits.MAX_ZIP_SINGLE_XML_BYTES
        ):
            raise ExtractionError("Office ZIP single XML member exceeds budget")
        if member.file_size > limits.ZIP_RATIO_MIN_UNCOMPRESSED_BYTES:
            if member.compress_size == 0:
                raise ExtractionError("Office ZIP compression ratio exceeds budget")
            if member.file_size / member.compress_size > limits.MAX_ZIP_COMPRESSION_RATIO:
                raise ExtractionError("Office ZIP compression ratio exceeds budget")


def _append_table_fragments(
    table: Table, table_number: int, collector: FragmentCollector
) -> None:
    seen_cells: set[object] = set()
    for row_number, row in enumerate(table.rows, 1):
        values: list[str] = []
        positions: list[int] = []
        for cell_number, cell in enumerate(row.cells, 1):
            element = cell._tc
            if element in seen_cells:
                continue
            seen_cells.add(element)
            value = cell.text.strip()
            if value:
                positions.append(cell_number)
                values.append(value)
        if values:
            locator = (
                f"table:{table_number};row:{row_number};"
                f"cells:{positions[0]}-{positions[-1]}"
            )
            collector.append(locator, "\t".join(values))


def _extract_docx_snapshot(path: Path) -> Extraction:
    _validate_office_zip(path)
    try:
        with path.open("rb") as stream:
            document = Document(stream)
    except Exception as exc:
        raise ExtractionError(f"failed to read DOCX file: {path}") from exc

    collector = FragmentCollector()
    paragraph_number = 0
    table_number = 0
    for item in document.iter_inner_content():
        if isinstance(item, Paragraph):
            paragraph_number += 1
            if item.text.strip():
                collector.append(f"paragraph:{paragraph_number}", item.text)
        elif isinstance(item, Table):
            table_number += 1
            _append_table_fragments(item, table_number, collector)
    return collector.extraction("extracted")


def extract_docx(path: Path) -> Extraction:
    with snapshot_file(path) as snapshot:
        return enforce_extraction_budget(_extract_docx_snapshot(snapshot))


def _extract_xlsx_snapshot(path: Path) -> Extraction:
    _validate_office_zip(path)
    stream = path.open("rb")
    try:
        book = load_workbook(
            stream,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        stream.close()
        raise ExtractionError(f"failed to read spreadsheet file: {path}") from exc

    try:
        collector = FragmentCollector()
        for sheet in book.worksheets:
            max_row = sheet.max_row or 0
            max_column = sheet.max_column or 0
            if (
                max_row > limits.MAX_WORKSHEET_ROWS
                or max_column > limits.MAX_WORKSHEET_COLUMNS
                or max_row * max_column > limits.MAX_WORKSHEET_CELLS
            ):
                raise ExtractionError(
                    f"worksheet dimensions exceed budget: {sheet.title} "
                    f"({max_row} rows x {max_column} columns)"
                )
        for sheet in book.worksheets:
            sheet.reset_dimensions()
        for sheet in book.worksheets:
            observed_max_column = 0
            for row_number, row in enumerate(sheet.iter_rows(), 1):
                row_width = len(row)
                observed_max_column = max(observed_max_column, row_width)
                if (
                    row_number > limits.MAX_WORKSHEET_ROWS
                    or row_width > limits.MAX_WORKSHEET_COLUMNS
                    or row_number * observed_max_column > limits.MAX_WORKSHEET_CELLS
                ):
                    raise ExtractionError(
                        f"worksheet dimensions exceed budget: {sheet.title} "
                        f"({row_number} observed rows x "
                        f"{observed_max_column} observed columns)"
                    )
                values = ["" if cell.value is None else str(cell.value) for cell in row]
                if any(values):
                    locator = (
                        f"sheet:{sheet.title};cells:A{row_number}-"
                        f"{get_column_letter(len(row))}{row_number}"
                    )
                    collector.append(locator, "\t".join(values))
        return collector.extraction("extracted")
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionError(f"failed to extract spreadsheet file: {path}") from exc
    finally:
        try:
            book.close()
        finally:
            stream.close()


def extract_xlsx(path: Path) -> Extraction:
    with snapshot_file(path) as snapshot:
        return enforce_extraction_budget(_extract_xlsx_snapshot(snapshot))


class DocxExtractor:
    suffixes = {".docx"}
    extract = staticmethod(extract_docx)
    extract_snapshot = staticmethod(_extract_docx_snapshot)


class XlsxExtractor:
    suffixes = {".xlsx", ".xlsm"}
    extract = staticmethod(extract_xlsx)
    extract_snapshot = staticmethod(_extract_xlsx_snapshot)


registry.register(DocxExtractor())
registry.register(XlsxExtractor())
